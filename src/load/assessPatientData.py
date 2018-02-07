'''
Created on 06/02/2018

@author: mark
'''
import os
from openpyxl import load_workbook

class LoadData: 
    def path(self,fil):
        
        pn=os.path.abspath(__file__)
        pn=pn.split("src")[0]
        

        #The data file path is now created where the data folder and dataFile.csv is referenced
        filepath=os.path.join(pn,'data')
        
        return filepath

    def loadFile(self,filepath):
        
        files=os.listdir(filepath)
        self.c_names=['N30','N34','N50']
        self.location=[]
        self.values={}
        for fs in files:
            
            fs=os.path.join(filepath,fs)
            
            #first, open the file 
            wb = load_workbook(fs,read_only=True)
            sNames=wb.sheetnames
        
            for s in sNames:
                ws=wb[s]
                col_indices = ws.rows
            
                i=0
                for c in col_indices:
                    if i==0:
                        n=0
                        for cc in c:
                            for ccc in self.c_names:
                                if cc.value==ccc:
                                    v=[]
                                    self.values[cc.value]=v
                                    self.location.append(n)
                                    
                            n+=1
                    
                    else:
                        for s in self.location:
                            for sl in self.values:
                                v=self.values[sl]
                                v.append(c[s].value)
                                
                            
                    i+=1
               
ld=LoadData()
filepath=ld.path('COL100_data_upd.xlsx')
ld.loadFile(filepath)
